#!/usr/bin/env python3
"""
Legacy Odyssey — Influencer CRM builder.
Generates a sortable, printable xlsx of every creator we've touched.
Run: python build-crm.py  ->  influencer-crm.xlsx
Re-run any time the roster changes; edit the ROWS list below.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- palette (brand) ----
INK   = "1A1510"
GOLD  = "C8A96E"
CREAM = "FAF7F2"
CARD  = "F0E8DC"
GREEN = "DDEBD8"
RED   = "F2D9D6"
YELLOW= "F6EFD6"

# status -> fill for quick scanning
STATUS_FILL = {
    "repeat":   GREEN,
    "posted":   GREEN,
    "accepted": GREEN,
    "contacted":YELLOW,
    "prospect": CARD,
    "rejected": RED,
    "ghosted":  RED,
}

# ---- columns ----
COLS = [
    ("Name", 20), ("Handle", 22), ("Platform", 16), ("Followers", 12),
    ("Found Via", 20), ("Niche", 26), ("US Aud %", 10), ("Eng %", 9),
    ("Price", 16), ("Rating", 8), ("Lane", 7), ("Status", 11),
    ("Last Contact", 13), ("Notes", 70),
]

# ---- the roster ----
# Each row: name, handle, platform, followers, found_via, niche, us_pct, eng_pct,
#           price, rating, lane, status, last_contact, notes
ROWS = [
    # ---------- WORKED WITH (real outcomes) ----------
    ["Giulia Busetto", "@giulibusetto_", "IG / TikTok", "33k IG", "Collabstr (mom/baby)",
     "Miami UGC mom; baby/lifestyle/travel", "no data", "24.7%", "$60 feed / $200 reel", 5, "A", "repeat",
     "2026-06-07",
     "Repeat collaborator (Dan: 'we already use Giulia'). Completed a $759 IG Reel (Jun). 5.0, Top Creator, Responds Fast. Best relationship to date. No audience-location data on profile but proven delivery + viral TikTok (2.2M avg views)."],

    ["Amanda Foust", "@amanda_arrows", "IG / TikTok / Amazon", "32.3k", "Collabstr (parenting)",
     "Parenting / mom lifestyle", "80%", "83.4%", "$145 feed", 5, "A", "contacted",
     "2026-06-16",
     "Order PLACED Jun 16 — IG Feed Post $159.50 (incl fee) funded from the Jayna refund balance. IN-PROGRESS, awaiting acceptance (~Jun 19). Top Creator. 87% F, 75% target age. Odessa FL. Highest-fit new pick: engagement is unicorn-level."],

    ["Akshita Thakur", "@rayanved", "IG", "15.8k", "Collabstr (mom/baby)",
     "Baby / motherhood; viral reels", "no data", "2.6%", "$50 story / $125 feed", 2, "A", "posted",
     "2026-06-03",
     "CLOSED — DO NOT PURSUE (Dan 2026-06-16): India-based, wrong market for US customer goals. Both items completed + paid out; she built her own book via comp code and posted. Warm/responsive but not a repeat. Rating reflects FIT for our US goals, not her professionalism. Do not surface her name/handle/child in any creative (facebook hard rule)."],

    ["Jayna Chambliss", "@_jblisss", "IG", "7.7k", "Collabstr (mom/baby)",
     "Mom / baby", "74%", "29.6%", "$250 feed", 1, "A", "ghosted",
     "2026-05-25",
     "Ghosted — order sent May 25, never accepted, auto-cancelled, $275 refunded (redeployed to Amanda Foust). 77% F. Do NOT re-order. Non-response is the answer."],

    ["Carina", "@a.calm.mom", "IG", "n/a", "Collabstr (mom/baby)",
     "Calm / gentle-parenting mom", "n/a", "n/a", "~$92/post (3 for $275)", 1, "A", "rejected",
     "2026-05-22",
     "Declined our order (May 22). No deep stats captured. Do not re-pursue without a reason."],

    ["Sarah Baraldi", "(Collabstr #105567)", "IG", "n/a", "Collabstr",
     "Baby/toddler sleep & motherhood", "n/a", "n/a", "$295 reel", 2, "A", "rejected",
     "2026-05-18",
     "She requested to cancel the order (May). Was a 4.8-rated creator (Woodbury NJ). Closed — remove from active pursuit unless she re-initiates."],

    # ---------- HELD / NO CLEAR PATH ----------
    ["Lianne", "@diaperdynasty", "IG", "n/a", "Collabstr",
     "Mom / diaper / baby", "n/a", "n/a", "n/a", 3, "A", "prospect",
     "2026-05",
     "ON HOLD per Dan (May 2026 — 'I don't want to use Lianne right now'). No deep stats captured. Re-evaluate only if Dan re-opens."],

    ["Meaghan Williamson", "@mindfullymademotherhood", "IG", "811k", "Her website + IG (mindfullymademotherhood.com)",
     "Mindful motherhood", "unknown", "unknown", "unknown (not on marketplaces)", 3, "B", "prospect",
     "—",
     "NOT on Collabstr or any marketplace; no public business email anywhere. Only outreach path = IG DM. Verify US audience before pursuing (US-first screen). Lane B: more realistic as organic/affiliate than a paid Collabstr post."],

    # ---------- ACTIVE PROSPECTS (Dan-sourced) ----------
    ["Hayley Carter", "@hayley__carter", "IG", "72.3k", "Dan-sourced / IG direct",
     "Motherhood / lifestyle / wellness; mom of 4; Dallas TX", "unknown (she's US)", "unknown", "unknown (agency-quoted)", 4, "A", "prospect",
     "—",
     "PROVISIONAL rating 4 — strong on visible signals (US-based Dallas creator, motherhood niche, public 'Digital creator', healthy 72k/1.7k follow ratio) but engagement % and audience US-geo are UNVERIFIED. AGENCY-REPPED: outreach goes to teamhayley@itgirlagency.com (email is the front door, not a cold DM). Bio: 'faith family fitness fashion, relatable mom content + wellness + ootd.' linktr.ee/HayleyCarter. Ask the agency for a media kit (engagement + audience geography) BEFORE committing budget; wants a permanent feed post (stays on grid)."],

    # ---------- TOP-10 SHORTLIST (researched 2026-06-10, not yet contacted) ----------
    ["Kristen Corrao", "@kristen_corrao", "IG / TikTok", "1.2k IG / 22.8k TT", "Collabstr (parenting)",
     "Motherhood / parenting", "88%", "18.7%", "$70 feed", 5, "B", "prospect",
     "2026-06-10 (researched)",
     "Excellent: 88% US, 71% F, 74% target age, Top Creator. Tiny IG following -> Lane B: recurring affiliate (35%) likely beats one-shot paid. If paid: cheapest high-fit option at $70."],

    ["Jennifer Mendez", "@jenn11", "IG", "2.2k", "Collabstr (parenting)",
     "Lifestyle / parenting", "98%", "2.0%", "$85 feed", 4, "B", "prospect",
     "2026-06-10 (researched)",
     "Highest US% of anyone (98%); 117k avg views (punches above follower count). 55% F (lower than ideal), age 25-34 52%. Fort Worth TX, 5.0 rating. Tiny following -> Lane B affiliate-first."],

    ["Denise Ferreira", "@deniiseferreira", "IG / TikTok", "12.4k", "Collabstr (mommy)",
     "Mommy / lifestyle", "85%", "0.4% IG", "$100 feed", 4, "A", "prospect",
     "2026-06-10 (researched)",
     "85% US, 96% F (highest), 68% target age. IG engagement low (0.4%) but TikTok viral (318k avg views). San Diego CA. Strong demo; value buy."],

    ["Alexandra 'Jax' Molieri", "@jaxdisneyfam", "IG", "16.7k", "Collabstr (motherhood)",
     "Family / Disney / UGC", "85%", "2.0%", "$75 story / $225 3-feed", 4, "A", "prospect",
     "2026-06-10 (researched)",
     "85% US, 85% F, 65% target age, 5.0 rating. Twentynine Palms CA. Clean family fit, fair pricing."],

    ["Audrey Murray", "@audreyinwanderlust_", "IG", "27.2k", "Collabstr (motherhood)",
     "Lifestyle / motherhood", "66%", "16.3%", "$400 feed (premium)", 4, "A", "prospect",
     "2026-06-10 (researched)",
     "Very high engagement (16.3%), 83% F, 66% target age, 4.9 rating. US only 66% and pricing is premium ($375 story / $400 feed / $750 reel). Clayton NC. Strong but expensive."],

    ["Keeping Up With The Griffins", "(Collabstr)", "IG / YouTube", "13.9k IG / 34.1k YT", "Collabstr (parenting)",
     "Family / lifestyle / comedy / travel", "78%", "1.7%", "$200 feed", 4, "A", "prospect",
     "2026-06-10 (researched)",
     "Most-proven on the list: 5.0 rating, 13 reviews, Top Creator, Responds Fast. 78% US, 66% F, age 25-34 47%. Charleston SC. Lower engagement but the track record de-risks it."],

    ["Kaci Costello", "@kieran.and.mama", "IG", "18.1k", "Collabstr (mommy)",
     "Mommy blogger", "54%", "n/a", "$100 feed", 3, "A", "prospect",
     "2026-06-10 (researched)",
     "Budget-friendly ($100 feed / $50 story), 71% F, 66% target age. US only 54% (under our 60% floor) — flagged by the US-first screen. Hopewell Junction NY."],

    ["Diana Capitao", "@capitao.diana", "IG", "14.3k", "Collabstr (parenting)",
     "Parenting expert", "56%", "0.7%", "$175 for 3 feed", 3, "A", "prospect",
     "2026-06-10 (researched)",
     "93% F (excellent), 63% target age, best $/post ($175 for 3). US only 56% (under floor) + low engagement (0.7%). Beverly MA."],

    ["Shelly Ray", "@shellyraycreates", "IG", "47.5k", "Collabstr (family)",
     "Family / lifestyle", "61%", "1.0%", "$450 feed (premium)", 3, "A", "prospect",
     "2026-06-10 (researched)",
     "Biggest following on the shortlist; 91% F, 67% target age, Top Creator, 5.0. US 61% (borderline), low engagement (1.0%), premium price ($450 feed / $650 reel). Opportunistic only."],
]

# ---- build workbook ----
wb = Workbook()
ws = wb.active
ws.title = "Influencer CRM"

thin = Side(style="thin", color="D8C9A8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor=INK)
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
wrap_top = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="top")

# header
for c, (title, width) in enumerate(COLS, start=1):
    cell = ws.cell(row=1, column=c, value=title)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(c)].width = width

# data
center_cols = {3,4,7,8,10,11,12,13}  # platform, followers, us%, eng%, rating, lane, status, last contact
for r, row in enumerate(ROWS, start=2):
    status = row[11]
    rowfill = STATUS_FILL.get(status, CREAM)
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        cell.alignment = center if c in center_cols else wrap_top
        cell.fill = PatternFill("solid", fgColor=rowfill)
        if c == 10:  # rating bold
            cell.font = Font(bold=True, color=INK)

# freeze header + name col; autofilter makes it sortable in Excel/Sheets
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(ROWS)+1}"
ws.row_dimensions[1].height = 30

# print setup — landscape, fit to one page wide, repeat header, gridlines
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_title_rows = "1:1"
ws.print_options.gridLines = True
ws.page_margins.left = ws.page_margins.right = 0.3
ws.page_margins.top = ws.page_margins.bottom = 0.5

# ---- rubric / legend sheet ----
ws2 = wb.create_sheet("Rubric & Legend")
ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 100
def hrow(ws, r, text):
    ws.cell(row=r, column=1, value=text).font = Font(bold=True, size=13, color=INK)
def kv(ws, r, k, v, kbold=True):
    a = ws.cell(row=r, column=1, value=k); a.font = Font(bold=kbold, color=INK); a.alignment = wrap_top
    b = ws.cell(row=r, column=2, value=v); b.alignment = wrap_top

r = 1
hrow(ws2, r, "Legacy Odyssey — Influencer CRM"); r += 1
kv(ws2, r, "Maintained by", "Influencer session. Re-run build-crm.py to regenerate after edits."); r += 2

hrow(ws2, r, "RATING RUBRIC (1–5) — fit for OUR US customer goals"); r += 1
kv(ws2, r, "5 — Ideal", "Buy / recruit now. >=75% US, >=80% female, target age 25-44 dominant, healthy engagement (>=3% IG or proven viral), proven (reviews / Top Creator / delivered for us), price sane for a $29 product."); r += 1
kv(ws2, r, "4 — Strong", "Pursue. Meets most criteria with ONE soft spot — e.g. US 60-75%, OR premium price, OR modest engagement offset by reach/demos."); r += 1
kv(ws2, r, "3 — Mixed", "Consider with caution. A real gap: US <60%, OR audience skews young/male, OR low engagement, OR unproven/no path."); r += 1
kv(ws2, r, "2 — Weak", "Opportunistic only. Multiple gaps OR a delivered-but-wrong-market creator (e.g. great work but non-US audience)."); r += 1
kv(ws2, r, "1 — Avoid", "Do not pursue. Disqualifying: mostly non-US audience, majority male, suspicious metrics, ghosted, or declined."); r += 2

hrow(ws2, r, "LANES"); r += 1
kv(ws2, r, "Lane A", "Regular poster. Paid Collabstr/marketplace posts on their own feed. Our default for mid-tier creators (~10k-50k) where one-shot cash buys real reach."); r += 1
kv(ws2, r, "Lane B", "Affiliate-only. Better as a Rewardful affiliate (35% recurring) + comp gift code than a one-shot paid post. Use for tiny-but-engaged creators and anyone with no marketplace presence. Coordinate with the affiliates session (they own the 19 Tier-A sub-10k targets)."); r += 2

hrow(ws2, r, "STATUS VALUES"); r += 1
for k, v in [
    ("prospect", "Identified/researched, not yet contacted."),
    ("contacted", "Order placed or DM sent; awaiting their response."),
    ("accepted", "Creator accepted the order/collab; work not yet delivered."),
    ("posted", "Content delivered and live."),
    ("repeat", "Delivered AND we've engaged them more than once / want to again."),
    ("rejected", "They declined or cancelled the order."),
    ("ghosted", "No response; order lapsed/auto-cancelled."),
]:
    kv(ws2, r, k, v, kbold=True); r += 1
r += 1

hrow(ws2, r, "STANDING RULES"); r += 1
kv(ws2, r, "US-first", "Screen creators for US base FIRST (Dan 2026-06-16). US-audience % matters MORE than raw engagement for our US customer goals."); r += 1
kv(ws2, r, "Names rule", "Do NOT put real outside people's names/handles or their children's names into any brand content/creative without Dan's explicit per-instance OK (facebook hard rule 2026-06-16)."); r += 1
kv(ws2, r, "Comp gift", "Prefer the /admin comp gift code (free year, digital — nothing ships) over cash where a creator will accept it."); r += 1
kv(ws2, r, "Word bans", "Never 'forever' / 'chapter' / 'family book/story'. It's a BABY book about the CHILD. Web editor is equal to the app (app not required)."); r += 1

# color key row
r += 1
hrow(ws2, r, "ROW COLOR KEY (CRM sheet)"); r += 1
for label, color in [("green = active/won (repeat/posted/accepted)", GREEN),
                     ("yellow = in flight (contacted)", YELLOW),
                     ("tan = prospect", CARD),
                     ("red = dead (rejected/ghosted)", RED)]:
    cell = ws2.cell(row=r, column=1, value=" ")
    cell.fill = PatternFill("solid", fgColor=color)
    ws2.cell(row=r, column=2, value=label).alignment = wrap_top
    r += 1

ws2.page_setup.orientation = "portrait"
ws2.sheet_properties.pageSetUpPr.fitToPage = True
ws2.page_setup.fitToWidth = 1

out = "influencer-crm.xlsx"
wb.save(out)
print("wrote", out, "—", len(ROWS), "creators")
